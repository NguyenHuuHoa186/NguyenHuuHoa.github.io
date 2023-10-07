from openpyxl import *
from tkinter import *
import re
from tkinter import ttk

wb = load_workbook('Book1.xlsx')

sheet = wb.active


def excel():
	

	sheet.column_dimensions['A'].width = 30
	sheet.column_dimensions['B'].width = 10
	sheet.column_dimensions['C'].width = 10
	sheet.column_dimensions['D'].width = 20
	sheet.column_dimensions['E'].width = 20
	sheet.column_dimensions['F'].width = 40
	sheet.column_dimensions['G'].width = 50


	
	sheet.cell(row=1, column=1).value = "MSSV"
	sheet.cell(row=1, column=2).value = "Họ tên"
	sheet.cell(row=1, column=3).value = "Ngày sinh"
	sheet.cell(row=1, column=4).value = "Email"
	sheet.cell(row=1, column=5).value = "SĐT"
	sheet.cell(row=1, column=6).value = "Học kì"
	sheet.cell(row=1, column=7).value = "Năm học"



def focus1(event):
	course_field.focus_set()


def focus2(event):
	
	sem_field.focus_set()



def focus3(event):
	form_no_field.focus_set()


def focus4(event):
	contact_no_field.focus_set()


def focus5(event):
	email_id_field.focus_set()


def focus6(event):
	address_field.focus_set()

def focus7(event):
	monhoc_field.focus_set()



def clear():
	
	name_field.delete(0, END)
	course_field.delete(0, END)
	sem_field.delete(0, END)
	form_no_field.delete(0, END)
	contact_no_field.delete(0, END)
	email_id_field.delete(0, END)
	address_field.delete(0, END)



def insert():
    if (name_field.get() == "" and
        course_field.get() == "" and
        sem_field.get() == "" and
        form_no_field.get() == "" and
        contact_no_field.get() == "" and
        email_id_field.get() == "" and
        address_field.get() == ""):
            
        print("Đã thêm thành công vào Excel")

    elif not name_field.get().isdigit() or len(name_field.get()) != 7:
        print("Mã số sinh viên phải là 7 chữ số")
        
    elif not re.match(r"[^@]+@[^@]+\.[^@]+", form_no_field.get()):
        print("Email không hợp lệ")
        
    elif not contact_no_field.get().isdigit() or len(contact_no_field.get()) != 10:
        print("Số điện thoại phải là 10 chữ số")

    elif email_id_field.get() not in ['1', '2', '3']:
        print("Học kỳ phải là 1, 2 hoặc 3")

    elif not re.match(r"^(0[1-9]|[12][0-9]|3[01])/(0[1-9]|1[012])/\d{4}$", sem_field.get()):
        print("Ngày sinh phải theo định dạng dd/mm/yyyy")

    else:
        current_row = sheet.max_row
        current_column = sheet.max_column

        sheet.cell(row=current_row + 1, column=1).value = name_field.get()
        sheet.cell(row=current_row + 1, column=2).value = course_field.get()
        sheet.cell(row=current_row + 1, column=3).value = sem_field.get()
        sheet.cell(row=current_row + 1, column=4).value = form_no_field.get()
        sheet.cell(row=current_row + 1, column=5).value = contact_no_field.get()
        sheet.cell(row=current_row + 1, column=6).value = email_id_field.get()
        sheet.cell(row=current_row + 1, column=7).value = address_field.get()


        wb.save('Book1.xlsx')

        name_field.focus_set()

        clear()


if __name__ == "__main__":
    
    root = Tk()

    root.configure(background='light green')

    root.title("Đăng ký học phần")

    root.geometry("500x300")

    excel()

    heading = Label(root, text="THÔNG TIN ĐĂNG KÝ HỌC PHẦN", bg="light green")

    name = Label(root, text="Mã số sinh viên", bg="light green")

    course = Label(root, text="Họ tên", bg="light green")

    sem = Label(root, text="Ngày sinh", bg="light green")

    form_no = Label(root, text="Email.", bg="light green")

    contact_no = Label(root, text="SĐT", bg="light green")

    email_id = Label(root, text="Học kì", bg="light green")

    address = Label(root, text="Năm học", bg="light green")

    monhoc = Label(root, text="Môn học", bg="light green")



    heading.grid(row=0, column=1)
    name.grid(row=1, column=0)
    course.grid(row=2, column=0)
    sem.grid(row=3, column=0)
    form_no.grid(row=4, column=0)
    contact_no.grid(row=5, column=0)
    email_id.grid(row=6, column=0)
    address.grid(row=7, column=0)
    monhoc.grid(row=8, column=0)



    name_field = Entry(root)
    course_field = Entry(root)
    sem_field = Entry(root)
    form_no_field = Entry(root)
    contact_no_field = Entry(root)
    email_id_field = Entry(root)

    
    name_field.bind("<Return>", focus1)

    
    course_field.bind("<Return>", focus2)

    
    sem_field.bind("<Return>", focus3)

    
    form_no_field.bind("<Return>", focus4)

    
    contact_no_field.bind("<Return>", focus5)


    email_id_field.bind("<Return>", focus6)


    
    name_field.grid(row=1, column=1, ipadx="100")
    course_field.grid(row=2, column=1, ipadx="100")
    sem_field.grid(row=3, column=1, ipadx="100")
    form_no_field.grid(row=4, column=1, ipadx="100")
    contact_no_field.grid(row=5, column=1, ipadx="100")
    email_id_field.grid(row=6, column=1, ipadx="100")


    academic_years = ['2022-2023', '2023-2024', '2024-2025']
    address_field = ttk.Combobox(root, values=academic_years)
    address_field.grid(row=7, column=1)

    excel()

    submit = Button(root, text="Đăng ký", fg="Black",
                            bg="Red", command=insert)
    submit.grid(row=12, column=1)

    exit_button = Button(root, text="Thoát", fg="Black",
                            bg="Red", command=root.quit)
    exit_button.grid(row=13, column=1)

    subjects = ["Lập trình python", "Lập trình java", "Công nghệ phần mềm", "Phát triển ứng dụng wed"]
    subject_vars = [IntVar() for _ in subjects]
    
    for i, subject in enumerate(subjects):
        Checkbutton(root, text=subject, variable=subject_vars[i]).grid(row=i+10, column=0, sticky=W)

    root.mainloop()
